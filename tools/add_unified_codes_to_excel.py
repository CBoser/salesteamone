"""
Add Unified Code column to IMPROVED_HOLT_BAT_NOVEMBER_2025.xlsm

This script adds a "Unified Code" column to all plan sheets that parses
the location codes from Column A into the standard format:
PPPP-PPP.000-EE-IIII

Example: "167010100 - 4085" → "1670-101.000-00-4085"
"""

import sys
sys.path.append('.')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from bat_system_v2.services.unified_code_parser import UnifiedCodeParser

SOURCE_FILE = "./BAT Files/IMPROVED_HOLT_BAT_NOVEMBER_2025.xlsm"
OUTPUT_FILE = "./BAT Files/IMPROVED_HOLT_BAT_WITH_CODES_NOVEMBER_2025.xlsm"

# Column positions
LOCATION_COL = 'A'  # Where raw codes are
UNIFIED_CODE_COL = 'AA'  # Where to put parsed unified codes (after existing columns)

# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
CODE_FONT = Font(name='Courier New', size=9)
CODE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def add_unified_code_column(ws, sheet_name, parser):
    """
    Add unified code column to a sheet

    Args:
        ws: Worksheet object
        sheet_name: Name of sheet
        parser: UnifiedCodeParser instance
    """
    print(f"    Processing: {sheet_name}")

    # Find data range
    max_row = ws.max_row
    if max_row < 5:
        print(f"      ⏭️  Skipped (no data)")
        return 0

    # Find header row (usually row 3 or 4)
    header_row = None
    for row in range(1, 6):
        cell_value = str(ws[f'{LOCATION_COL}{row}'].value or '').upper()
        if 'LOCATION' in cell_value or 'CODE' in cell_value or row == 4:
            header_row = row
            break

    if not header_row:
        header_row = 4  # Default

    # Add header
    header_cell = ws[f'{UNIFIED_CODE_COL}{header_row}']
    header_cell.value = "UNIFIED CODE"
    header_cell.font = HEADER_FONT
    header_cell.fill = HEADER_FILL
    header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Note: Comments removed to avoid VBA conflicts in macro-enabled files
    # See CODE LEGEND sheet for code explanation

    # Process data rows
    codes_added = 0
    for row in range(header_row + 1, max_row + 1):
        location_cell = ws[f'{LOCATION_COL}{row}']
        location_value = location_cell.value

        if not location_value:
            continue

        # Try to parse the location code
        try:
            raw_code = str(location_value).strip()

            # Skip if it's just a header or label
            if raw_code.upper() in ['LOCATION', 'CODE', '', 'N/A']:
                continue

            # Parse the code
            parsed = parser.parse(raw_code)

            if parsed.is_valid:
                # Add parsed code to column
                code_cell = ws[f'{UNIFIED_CODE_COL}{row}']
                code_cell.value = parsed.full_code
                code_cell.font = CODE_FONT
                code_cell.fill = CODE_FILL
                code_cell.alignment = Alignment(horizontal='left')
                codes_added += 1

        except Exception as e:
            # Skip problematic cells
            print(f"      ⚠️  Row {row}: {e}")
            continue

    # Set column width
    ws.column_dimensions[UNIFIED_CODE_COL].width = 22

    print(f"      ✅ Added {codes_added} unified codes")
    return codes_added


def add_code_legend_sheet(wb, parser):
    """
    Add a "CODE LEGEND" sheet explaining the unified code system

    Args:
        wb: Workbook object
        parser: UnifiedCodeParser instance
    """
    print("\n📖 Adding CODE LEGEND sheet...")

    # Create new sheet
    if "📋 CODE LEGEND" in wb.sheetnames:
        wb.remove(wb["📋 CODE LEGEND"])

    ws = wb.create_sheet("📋 CODE LEGEND", 0)

    # Title
    ws['A1'] = "UNIFIED CODE SYSTEM - Quick Reference"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:F1')

    # Format explanation
    row = 3
    ws[f'A{row}'] = "CODE FORMAT:"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    ws[f'A{row}'] = "PPPP-PPP.000-EE-IIII"
    ws[f'A{row}'].font = Font(name='Courier New', size=14, bold=True, color="0070C0")

    row += 2
    ws[f'A{row}'] = "EXAMPLE:"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    ws[f'A{row}'] = "1670-101.000-AB-4085"
    ws[f'A{row}'].font = Font(name='Courier New', size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Component breakdown
    row += 2
    ws[f'A{row}'] = "COMPONENT BREAKDOWN:"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:F{row}')

    # Headers
    row += 1
    headers = ["Component", "Value", "Length", "Description", "Example", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    # Components
    components = [
        ("PPPP", "1670", "4 digits", "Plan Number", "1670, 2336, 0383", "Left-padded with zeros"),
        ("PPP", "101", "3 digits", "Phase/Pack Code", "101=Foundation, 200=Walls", "See pack codes below"),
        ("000", "000", "3 digits", "Padding", "Always 000", "Reserved for future use"),
        ("EE", "AB", "2 chars", "Elevation Code", "AB, CD, 00", "A/B/C/D or numeric"),
        ("IIII", "4085", "4 digits", "Item Type", "4085=Fasteners", "See categories below"),
    ]

    for component in components:
        row += 1
        for col_idx, value in enumerate(component, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = value
            if col_idx == 1:
                cell.font = Font(name='Courier New', bold=True)
            elif col_idx == 2:
                cell.font = Font(name='Courier New')
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Pack codes
    row += 3
    ws[f'A{row}'] = "COMMON PACK CODES:"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    ws[f'A{row}'] = "Code"
    ws[f'B{row}'] = "Pack Name"
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL

    pack_codes = [
        ("101", "FOUNDATION"),
        ("102", "MAIN FLOOR SYSTEM"),
        ("200", "MAIN WALLS"),
        ("201", "PARTITION WALLS"),
        ("300", "ROOF FRAMING"),
        ("400", "EXTERIOR TRIM"),
        ("500", "INTERIOR TRIM"),
        ("05", "UNDERFLOOR"),
        ("10", "FOUNDATION (alt)"),
        ("20", "MAIN WALLS (alt)"),
    ]

    for code, name in pack_codes:
        row += 1
        ws[f'A{row}'] = code
        ws[f'B{row}'] = name
        ws[f'A{row}'].font = Font(name='Courier New')

    # Item type categories
    row += 3
    ws[f'A{row}'] = "ITEM TYPE CATEGORIES:"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    row += 1
    ws[f'A{row}'] = "Range"
    ws[f'B{row}'] = "Category"
    ws[f'C{row}'] = "Examples"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL

    categories = [
        ("1000-1999", "Lumber (structural)", "Beams, headers, joists"),
        ("2000-2999", "Lumber (framing)", "Studs, plates"),
        ("3000-3999", "Lumber (trim)", "Trim boards, molding"),
        ("4000-4999", "Hardware/Fasteners", "Nails, bolts, anchors"),
        ("5000-5999", "Panels/Sheathing", "Plywood, OSB"),
        ("6000-6999", "Doors/Windows", "Entry doors, windows"),
        ("7000-7999", "Roofing", "Shingles, felt, vents"),
        ("8000-8999", "Misc Materials", "Insulation, caulk"),
        ("9000-9999", "Unclassified", "Default category"),
    ]

    for range_str, category, examples in categories:
        row += 1
        ws[f'A{row}'] = range_str
        ws[f'B{row}'] = category
        ws[f'C{row}'] = examples
        ws[f'A{row}'].font = Font(name='Courier New')

    # Usage examples
    row += 3
    ws[f'A{row}'] = "USAGE EXAMPLES:"
    ws[f'A{row}'].font = Font(bold=True, size=12)

    examples = [
        ("1670-101.000-00-4085", "Fasteners for Plan 1670 Foundation (all elevations)"),
        ("2336-200.000-AB-2085", "Studs for Plan 2336 Main Walls (elevations A & B)"),
        ("0383-300.000-CD-3085", "Roof trim for Plan 383 (elevations C & D)"),
    ]

    for code, description in examples:
        row += 1
        ws[f'A{row}'] = code
        ws[f'B{row}'] = description
        ws[f'A{row}'].font = Font(name='Courier New', bold=True)
        ws[f'A{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25

    print("   ✅ CODE LEGEND sheet added")


def main():
    """Main processing function"""
    print("=" * 70)
    print("Adding Unified Codes to HOLT BAT")
    print("=" * 70)
    print(f"\nSource: {SOURCE_FILE}")
    print(f"Output: {OUTPUT_FILE}\n")

    # Load workbook
    print("📖 Loading workbook...")
    wb = load_workbook(SOURCE_FILE, keep_vba=True)
    print(f"   ✅ Loaded {len(wb.sheetnames)} sheets\n")

    # Create parser
    parser = UnifiedCodeParser()

    # Add code legend sheet first
    add_code_legend_sheet(wb, parser)

    # Process plan sheets
    print("\n🔧 Adding unified code columns to plan sheets...")
    total_codes = 0
    sheets_processed = 0

    # Skip documentation sheets
    skip_sheets = ['📖 INSTRUCTIONS', '📐 FORMULA GUIDE', '💡 SUGGESTIONS', '📋 CODE LEGEND',
                   'PriceData', 'Summary', 'Index']

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue

        ws = wb[sheet_name]
        codes_added = add_unified_code_column(ws, sheet_name, parser)

        if codes_added > 0:
            total_codes += codes_added
            sheets_processed += 1

    # Save workbook
    print(f"\n💾 Saving workbook...")
    wb.save(OUTPUT_FILE)

    # Summary
    print("\n" + "=" * 70)
    print("📊 SUMMARY")
    print("=" * 70)
    print(f"Sheets processed: {sheets_processed}")
    print(f"Total unified codes added: {total_codes:,}")
    print(f"Code legend sheet: Added")
    print(f"\n✅ Success!")
    print("=" * 70)
    print(f"\n📂 LOCATION")
    print("=" * 70)
    print(f"{OUTPUT_FILE}")
    print("\n" + "=" * 70)
    print("📖 NEXT STEPS")
    print("=" * 70)
    print("1. Open the new file")
    print("2. Check the '📋 CODE LEGEND' sheet for code explanation")
    print("3. Look for the 'UNIFIED CODE' column (Column AA) in plan sheets")
    print("4. Codes are automatically parsed from Location column")
    print("5. Use codes for filtering, sorting, and reporting")
    print("\n🎉 Unified codes added successfully!")
    print("=" * 70)


if __name__ == "__main__":
    main()
