#!/usr/bin/env python3
"""
Create Excel Improvement Guide
For use alongside existing HOLT BAT NOVEMBER 2025 .xlsm file

This creates a standalone guide with all formula improvements,
documentation, and step-by-step instructions that can be applied
to the existing macro-enabled file manually.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
import sys
import os

# Add create_improved_bat to path
sys.path.insert(0, os.path.dirname(__file__))
from create_improved_bat import (
    create_documentation_sheet,
    create_formula_guide_sheet,
    create_suggestions_sheet,
    COLORS
)

OUTPUT_FILE = "./BAT Files/HOLT_BAT_IMPROVEMENT_GUIDE.xlsx"

def create_manual_improvement_sheet(wb):
    """Create step-by-step manual improvement instructions"""
    ws = wb.create_sheet("🔧 MANUAL IMPROVEMENTS", 0)

    # Title
    ws['A1'] = "Step-by-Step Improvement Guide"
    ws['A1'].font = Font(size=18, bold=True, color="1F4E78")
    ws.merge_cells('A1:F1')

    ws['A2'] = "Use this guide to manually improve your HOLT BAT NOVEMBER 2025.xlsm file"
    ws['A2'].font = Font(size=11, italic=True)
    ws.merge_cells('A2:F2')

    # Ste instructions
    row = 4
    ws[f'A{row}'] = "📋 HOW TO USE THIS GUIDE"
    ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')

    instructions = [
        "1. Open your HOLT BAT NOVEMBER 2025.xlsm file in Excel",
        "2. Open this guide file side-by-side (View → Arrange All → Vertical)",
        "3. Follow the step-by-step improvements below",
        "4. Test each change before moving to the next",
        "5. Save a backup before making changes!",
        "",
        "⚠️ IMPORTANT: Make a backup copy of your file before starting!",
    ]

    row += 2
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        if instruction.startswith("⚠️"):
            ws[f'A{row}'].fill = PatternFill(start_color=COLORS['warning'], end_color=COLORS['warning'], fill_type='solid')
            ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

    # Step 1: Fix Margin% formula
    row += 2
    ws[f'A{row}'] = "STEP 1: Fix Margin% Formula (Division-by-Zero Protection)"
    ws[f'A{row}'].font = Font(size=13, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "Current Problem:"
    ws[f'A{row}'].font = Font(bold=True, color="C00000")
    row += 1
    ws[f'A{row}'] = "Column R (MARGIN%) formula crashes with #DIV/0! error when Total Sell (Column M) is zero"
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "OLD FORMULA (Column R):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = '=1-(P12/M12)'
    ws[f'A{row}'].font = Font(size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['error'], end_color=COLORS['error'], fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "NEW FORMULA (Column R):"
    ws[f'A{row}'].font = Font(bold=True, color="008000")
    row += 1
    ws[f'A{row}'] = '=IF(M12>0, 1-(P12/M12), 0)'
    ws[f'A{row}'].font = Font(size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "How to Apply:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    steps = [
        "1. In your HOLT BAT file, go to any plan sheet (e.g., 2336-B)",
        "2. Click on cell R5 (first data row in MARGIN% column)",
        "3. Look at the formula bar - you'll see: =1-(P5/M5)",
        "4. Click in the formula bar and change it to: =IF(M5>0, 1-(P5/M5), 0)",
        "5. Press Enter",
        "6. Copy this cell (Ctrl+C)",
        "7. Select the entire Column R from row 5 down to your last data row",
        "8. Paste (Ctrl+V) to apply to all rows",
        "9. Repeat for all plan sheets",
    ]

    for step in steps:
        ws[f'A{row}'] = step
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

    row += 1
    ws[f'A{row}'] = "✅ Benefit: No more #DIV/0! errors - formula returns 0 instead of crashing"
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')

    # Step 2: Add IFERROR to VLOOKUP formulas
    row += 3
    ws[f'A{row}'] = "STEP 2: Add Error Handling to Price Lookups"
    ws[f'A{row}'].font = Font(size=13, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "Current Problem:"
    ws[f'A{row}'].font = Font(bold=True, color="C00000")
    row += 1
    ws[f'A{row}'] = "Column H (PRICE) shows #N/A error when SKU is not found in PriceData"
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "OLD FORMULA (Column H - simplified example):"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    ws[f'A{row}'] = '=VLOOKUP(F12,PriceData,17,0)'
    ws[f'A{row}'].font = Font(size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['error'], end_color=COLORS['error'], fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "NEW FORMULA (Column H):"
    ws[f'A{row}'].font = Font(bold=True, color="008000")
    row += 1
    ws[f'A{row}'] = '=IFERROR(VLOOKUP(F12,PriceData,17,0), 0)'
    ws[f'A{row}'].font = Font(size=11)
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "How to Apply:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    steps2 = [
        "1. In your HOLT BAT file, go to any plan sheet",
        "2. Click on cell H5 (first data row in PRICE column)",
        "3. Look at the existing formula - it will have VLOOKUP or INDEX-MATCH",
        "4. Wrap the entire formula in IFERROR(..., 0)",
        "   Before: =VLOOKUP(...)",
        "   After:  =IFERROR(VLOOKUP(...), 0)",
        "5. Copy this cell and paste to all rows in Column H",
        "6. Repeat for all plan sheets",
    ]

    for step in steps2:
        ws[f'A{row}'] = step
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        if step.startswith("   "):
            ws[f'A{row}'].font = Font(size=10)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

    row += 1
    ws[f'A{row}'] = "✅ Benefit: No more #N/A errors - shows $0 instead when SKU not found"
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')

    # Step 3: Add Data Validation
    row += 3
    ws[f'A{row}'] = "STEP 3: Add Price Level Dropdown (Data Validation)"
    ws[f'A{row}'].font = Font(size=13, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "Current Problem:"
    ws[f'A{row}'].font = Font(bold=True, color="C00000")
    row += 1
    ws[f'A{row}'] = "Column S (Price Level) allows any text - can type invalid values like '05' or 'ABC'"
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    ws[f'A{row}'] = "Solution: Add dropdown list"
    ws[f'A{row}'].font = Font(bold=True, color="008000")

    row += 2
    ws[f'A{row}'] = "How to Apply:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    steps3 = [
        "1. In your HOLT BAT file, go to any plan sheet",
        "2. Select Column S (entire column or just data rows S5:S1000)",
        "3. Go to Data tab → Data Validation",
        "4. In 'Allow' dropdown, select 'List'",
        "5. In 'Source' box, type: 01,02,03,L5",
        "6. Check 'Show dropdown' option",
        "7. Click OK",
        "8. Repeat for all plan sheets",
    ]

    for step in steps3:
        ws[f'A{row}'] = step
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

    row += 1
    ws[f'A{row}'] = "✅ Benefit: Dropdown prevents typos - can only select valid price levels"
    ws[f'A{row}'].fill = PatternFill(start_color=COLORS['success'], end_color=COLORS['success'], fill_type='solid')
    ws[f'A{row}'].font = Font(bold=True)
    ws.merge_cells(f'A{row}:F{row}')

    # Quick wins
    row += 3
    ws[f'A{row}'] = "QUICK WINS (Optional but Recommended)"
    ws[f'A{row}'].font = Font(size=13, bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type='solid')
    ws.merge_cells(f'A{row}:F{row}')

    row += 2
    quick_wins = [
        ("Freeze Panes", "View → Freeze Panes → Freeze Top Row", "Keep headers visible when scrolling"),
        ("Protect Formulas", "Select formula columns → Format Cells → Protection → Locked", "Prevent accidental deletion"),
        ("Color Code", "Yellow fill for calculated columns (H, M, P, Q, R)", "Visual guide"),
        ("Add Comments", "Right-click column headers → Insert Comment → Add notes", "Help for future users"),
        ("Named Ranges", "Formulas → Define Name → Create PriceData range", "Easier to read formulas"),
    ]

    for title, how, benefit in quick_wins:
        ws[f'A{row}'] = f"• {title}:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = f"  How: {how}"
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        ws[f'A{row}'] = f"  Benefit: {benefit}"
        ws[f'A{row}'].font = Font(italic=True)
        ws.merge_cells(f'A{row}:F{row}')
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 100

    return ws


def create_formula_examples_sheet(wb):
    """Create sheet with copyable formula examples"""
    ws = wb.create_sheet("📋 FORMULA EXAMPLES")

    # Title
    ws['A1'] = "Copyable Formula Examples"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:D1')

    ws['A2'] = "Copy these formulas directly into your HOLT BAT file"
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:D2')

    # Headers
    row = 4
    headers = ['Column', 'Formula Purpose', 'Copy This Formula', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type='solid')

    # Formula examples
    formulas = [
        ('R', 'Margin % (protected)', '=IF(M5>0, 1-(P5/M5), 0)', 'Replace row 5 with your starting row'),
        ('H', 'Price with error handling', '=IFERROR(VLOOKUP(F5,PriceData,17,0), 0)', 'Wrap existing VLOOKUP'),
        ('M', 'Total Sell', '=IF(G5>0, H5*G5, "")', 'Returns blank if no quantity'),
        ('P', 'Total Cost', '=IF(G5>0, V5*G5, "")', 'Returns blank if no quantity'),
        ('Q', 'Margin Dollars', '=IF(M5>0, M5-P5, "")', 'Profit in dollars'),
    ]

    row += 1
    for col_name, purpose, formula, notes in formulas:
        ws.cell(row=row, column=1).value = col_name
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).value = purpose
        ws.cell(row=row, column=3).value = formula
        ws.cell(row=row, column=3).font = Font(size=10)
        ws.cell(row=row, column=3).fill = PatternFill(start_color=COLORS['calculated'], end_color=COLORS['calculated'], fill_type='solid')
        ws.cell(row=row, column=4).value = notes
        ws.cell(row=row, column=4).font = Font(italic=True, size=9)
        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 30

    return ws


def main():
    """Create Excel improvement guide"""
    print("Creating Excel Improvement Guide...")
    print(f"Output: {OUTPUT_FILE}\n")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets
    print("📝 Creating guide sheets...")
    create_manual_improvement_sheet(wb)
    create_formula_examples_sheet(wb)
    create_documentation_sheet(wb)
    create_formula_guide_sheet(wb)
    create_suggestions_sheet(wb)

    print("✅ Created 5 guide sheets\n")

    # Save
    print(f"💾 Saving to {OUTPUT_FILE}...")
    try:
        os.makedirs(os.path.dirname(OUTPUT_FILE) or '.', exist_ok=True)
        wb.save(OUTPUT_FILE)
        print(f"✅ Success!\n")
        print(f"📂 Location: {os.path.abspath(OUTPUT_FILE)}")
        print(f"\n📖 How to use:")
        print(f"1. Open your HOLT BAT NOVEMBER 2025.xlsm file")
        print(f"2. Open this guide file: {OUTPUT_FILE}")
        print(f"3. Arrange windows side-by-side (View → Arrange All → Vertical)")
        print(f"4. Follow the step-by-step instructions in the first sheet")
        print(f"5. Copy formulas from the FORMULA EXAMPLES sheet as needed")
        print(f"\n⚠️  Make a backup of your HOLT BAT file before making changes!")
        print(f"\n🎉 Ready to improve your Excel file!")
    except Exception as e:
        print(f"❌ Error saving: {e}")


if __name__ == "__main__":
    main()
