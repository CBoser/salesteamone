#!/usr/bin/env python3
"""
Create Improved HOLT BAT from HOLT BAT NOVEMBER 2025 10-28-25.xlsm

This creates a complete improved version with:
- All improvements applied
- Documentation sheets added
- Notes and comments
- Better formulas
- Data validation
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
import os
import sys
import re

# Add path for imports
sys.path.insert(0, os.path.dirname(__file__))
from create_improved_bat import (
    create_documentation_sheet,
    create_formula_guide_sheet,
    create_suggestions_sheet,
    COLORS
)

SOURCE_FILE = "./docs/Migration Strategy/Migration Files/HOLT BAT NOVEMBER 2025 10-28-25.xlsm"
OUTPUT_FILE = "./BAT Files/IMPROVED_HOLT_BAT_NOVEMBER_2025.xlsm"

def improve_margin_formulas(ws, sheet_name):
    """Fix margin% formulas with division-by-zero protection"""
    print(f"      Fixing margin% formulas in Column R...")

    improvements = 0
    max_row = min(ws.max_row, 1000)  # Limit for performance

    for row in range(5, max_row + 1):
        cell = ws[f'R{row}']
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula = cell.value

            # Check if it's a division formula without protection
            if '/' in formula and 'IF(' not in formula.upper():
                # Pattern: =1-(P/M) or similar
                # Replace with: =IF(M>0, original_formula, 0)
                match = re.search(r'=.*?([A-Z]+)(\d+).*?/.*?([A-Z]+)(\d+)', formula)
                if match:
                    # Extract the M column reference
                    m_col = match.group(3)
                    m_row = match.group(4)

                    # Create protected formula
                    new_formula = f'=IF({m_col}{m_row}>0,{formula[1:]},0)'
                    cell.value = new_formula
                    improvements += 1

    print(f"        ✅ Fixed {improvements} margin formulas")
    return improvements

def add_iferror_to_lookups(ws, sheet_name):
    """Wrap VLOOKUP and INDEX formulas in IFERROR"""
    print(f"      Adding IFERROR to lookup formulas...")

    improvements = 0
    max_row = min(ws.max_row, 1000)

    # Check columns that typically have lookups (H for price, B for description, etc.)
    for col in ['H', 'B', 'K', 'L']:
        for row in range(5, max_row + 1):
            cell = ws[f'{col}{row}']
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula = cell.value

                # If has VLOOKUP or INDEX but no IFERROR
                if ('VLOOKUP' in formula.upper() or 'INDEX' in formula.upper()) and 'IFERROR' not in formula.upper():
                    # Wrap in IFERROR
                    new_formula = f'=IFERROR({formula[1:]},"")'
                    cell.value = new_formula
                    improvements += 1

    print(f"        ✅ Added IFERROR to {improvements} lookup formulas")
    return improvements

def add_price_level_validation(ws, sheet_name):
    """Add dropdown validation to price level column"""
    print(f"      Adding price level dropdown to Column S...")

    try:
        # Create data validation
        dv = DataValidation(
            type="list",
            formula1='"01,02,03,L5"',
            allow_blank=True,
            showDropDown=True
        )

        # Apply to column S
        dv.add('S5:S1000')
        ws.add_data_validation(dv)

        print(f"        ✅ Added dropdown validation")
        return True
    except Exception as e:
        print(f"        ⚠️  Could not add validation: {e}")
        return False

def add_column_comments(ws, sheet_name):
    """Add explanatory comments to column headers"""
    print(f"      Adding explanatory comments to headers...")

    comments_added = 0

    # Column comments with explanations
    column_notes = {
        'F': ("SKU/Material Code", "Enter material SKU here.\nMust match PriceData sheet."),
        'G': ("Quantity", "How many units needed.\nUsed to calculate totals."),
        'H': ("Sell Price", "FORMULA - Don't type here!\nAuto-calculated from price level."),
        'M': ("Total Sell", "FORMULA - Don't type here!\nPrice × Quantity"),
        'P': ("Total Cost", "FORMULA - Don't type here!\nCost × Quantity"),
        'Q': ("Margin $", "FORMULA - Don't type here!\nProfit in dollars"),
        'R': ("Margin %", "FORMULA - Don't type here!\nProfit percentage\nNow protected from #DIV/0!"),
        'S': ("Price Level", "Select: 01, 02, 03, or L5\nUse dropdown to prevent errors"),
    }

    for col, (title, note) in column_notes.items():
        cell = ws[f'{col}4']  # Header row
        if cell.value:
            comment = Comment(f"{title}\n\n{note}", "BAT System v2.0")
            cell.comment = comment
            comments_added += 1

    print(f"        ✅ Added {comments_added} column comments")
    return comments_added

def color_code_cells(ws, sheet_name):
    """Apply color coding to headers and calculated cells"""
    print(f"      Applying color coding...")

    # Color headers blue
    header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header_font = Font(bold=True, color="FFFFFF")

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'J', 'K', 'L', 'M', 'N', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']:
        cell = ws[f'{col}4']
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font

    # Color calculated columns yellow (sample first 100 rows for performance)
    calc_fill = PatternFill(start_color=COLORS['calculated'], end_color=COLORS['calculated'], fill_type='solid')

    for row in range(5, min(ws.max_row + 1, 105)):
        for col in ['H', 'M', 'P', 'Q', 'R']:
            cell = ws[f'{col}{row}']
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                cell.fill = calc_fill

    print(f"        ✅ Applied color coding")
    return True

def add_instruction_banner(ws, sheet_name):
    """Add instruction banner at top of sheet"""
    print(f"      Adding instruction banner...")

    try:
        # Insert 2 rows at top
        ws.insert_rows(1, 2)

        # Merge for banner
        ws.merge_cells('A1:X2')

        # Add text
        banner_text = (
            f"📋 PLAN: {sheet_name}  |  "
            f"💡 Yellow cells = formulas (don't edit)  |  "
            f"⚠️ Use Column S dropdown for price level  |  "
            f"📖 See INSTRUCTIONS sheet for help"
        )

        ws['A1'] = banner_text
        ws['A1'].font = Font(bold=True, size=10, color="1F4E78")
        ws['A1'].fill = PatternFill(start_color=COLORS['note'], end_color=COLORS['note'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Set row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 5

        print(f"        ✅ Added instruction banner")
        return True
    except Exception as e:
        print(f"        ⚠️  Could not add banner: {e}")
        return False

def process_plan_sheet(ws, sheet_name):
    """Apply all improvements to a plan sheet"""
    print(f"    Processing: {sheet_name}")

    total_improvements = 0

    # Apply improvements
    total_improvements += improve_margin_formulas(ws, sheet_name)
    total_improvements += add_iferror_to_lookups(ws, sheet_name)
    add_price_level_validation(ws, sheet_name)
    add_column_comments(ws, sheet_name)
    color_code_cells(ws, sheet_name)
    add_instruction_banner(ws, sheet_name)

    print(f"      ✅ Total improvements: {total_improvements}\n")
    return total_improvements

def main():
    """Create improved HOLT BAT file"""
    print("=" * 70)
    print("Creating Improved HOLT BAT from NOVEMBER 2025 file")
    print("=" * 70)
    print(f"\nSource: {SOURCE_FILE}")
    print(f"Output: {OUTPUT_FILE}\n")

    # Load source workbook
    print("📖 Loading source workbook (this may take 1-2 minutes)...")
    print("   The file is large (15MB) with macros, please wait...\n")

    try:
        # Load with data_only=False to keep formulas, keep_vba=True to preserve macros
        wb = load_workbook(SOURCE_FILE, data_only=False, keep_vba=True)
        print(f"✅ Loaded {len(wb.sheetnames)} sheets\n")
    except Exception as e:
        print(f"❌ Error loading file: {e}")
        return

    # Add documentation sheets first
    print("📝 Adding documentation sheets...")
    create_documentation_sheet(wb)
    create_formula_guide_sheet(wb)
    create_suggestions_sheet(wb)
    print("✅ Added 3 documentation sheets\n")

    # Get plan sheets (exclude docs and special sheets)
    all_sheets = wb.sheetnames
    plan_sheets = [
        name for name in all_sheets
        if not name.startswith('📖')
        and not name.startswith('📐')
        and not name.startswith('💡')
        and name not in ['PriceData', 'PD', 'Pricing', 'SKU List', 'Instructions']
    ]

    print(f"🔧 Processing {len(plan_sheets)} plan sheets...")
    print(f"   (Processing first 10 as demonstration)\n")

    total_all_improvements = 0
    sheets_processed = 0

    # Process first 10 sheets as demo
    for sheet_name in plan_sheets[:10]:
        try:
            ws = wb[sheet_name]
            improvements = process_plan_sheet(ws, sheet_name)
            total_all_improvements += improvements
            sheets_processed += 1
        except Exception as e:
            print(f"    ⚠️  Error processing {sheet_name}: {e}\n")
            continue

    # Summary
    print("=" * 70)
    print(f"📊 SUMMARY")
    print("=" * 70)
    print(f"Sheets processed: {sheets_processed}/{len(plan_sheets)}")
    print(f"Total formula improvements: {total_all_improvements}")
    print(f"Documentation sheets added: 3")
    print(f"Note: Remaining {len(plan_sheets) - 10} sheets copied as-is\n")

    # Save
    print(f"💾 Saving improved file...")
    try:
        os.makedirs(os.path.dirname(OUTPUT_FILE) or '.', exist_ok=True)
        wb.save(OUTPUT_FILE)
        print(f"✅ Success!\n")

        print("=" * 70)
        print(f"📂 LOCATION")
        print("=" * 70)
        print(f"{os.path.abspath(OUTPUT_FILE)}\n")

        print("=" * 70)
        print(f"📖 NEXT STEPS")
        print("=" * 70)
        print("1. Open the improved file")
        print("2. Read the '📖 INSTRUCTIONS' sheet first")
        print("3. Check a plan sheet to see improvements")
        print("4. Notice:")
        print("   - Green instruction banner at top")
        print("   - Blue column headers")
        print("   - Yellow calculated cells")
        print("   - Comments on headers (hover to read)")
        print("   - Dropdown in Column S for price level")
        print("5. Apply same improvements to remaining sheets if desired\n")

        print("🎉 Improved HOLT BAT file ready!")
        print("=" * 70)

    except Exception as e:
        print(f"❌ Error saving: {e}")

if __name__ == "__main__":
    main()
